###################################################################################################
# Cerebri AI CONFIDENTIAL
# Copyright (c) 2017-2020 Cerebri AI Inc., All Rights Reserved.
#
# NOTICE: All information contained herein is, and remains the property of Cerebri AI Inc.
# and its subsidiaries, including Cerebri AI Corporation (together “Cerebri AI”).
# The intellectual and technical concepts contained herein are proprietary to Cerebri AI
# and may be covered by U.S., Canadian and Foreign Patents, patents in process, and are
# protected by trade secret or copyright law.
# Dissemination of this information or reproduction of this material is strictly
# forbidden unless prior written permission is obtained from Cerebri AI. Access to the
# source code contained herein is hereby forbidden to anyone except current Cerebri AI
# employees or contractors who have executed Confidentiality and Non-disclosure agreements
# explicitly covering such access.
#
# The copyright notice above does not evidence any actual or intended publication or
# disclosure of this source code, which includes information that is confidential and/or
# proprietary, and is a trade secret, of Cerebri AI. ANY REPRODUCTION, MODIFICATION,
# DISTRIBUTION, PUBLIC PERFORMANCE, OR PUBLIC DISPLAY OF OR THROUGH USE OF THIS SOURCE
# CODE WITHOUT THE EXPRESS WRITTEN CONSENT OF CEREBRI AI IS STRICTLY PROHIBITED, AND IN
# VIOLATION OF APPLICABLE LAWS AND INTERNATIONAL TREATIES. THE RECEIPT OR POSSESSION OF
# THIS SOURCE CODE AND/OR RELATED INFORMATION DOES NOT CONVEY OR IMPLY ANY RIGHTS TO
# REPRODUCE, DISCLOSE OR DISTRIBUTE ITS CONTENTS, OR TO MANUFACTURE, USE, OR SELL
# ANYTHING THAT IT MAY DESCRIBE, IN WHOLE OR IN PART.
###################################################################################################
#!/home/dev/.conda/envs/py365/bin/python3.6
##########################################################################################################
## LOCATION
##########################################################################################################
import pickle
from pyspark.sql import SparkSession
from pyspark.sql.functions import col
from xlsxwriter.workbook import Workbook
import pandas as pd
from openpyxl import load_workbook, Workbook
import importlib.util
from pyspark.sql.functions import lit
import os
import config_te_reporting as cfg 
import sys

def start_spark_session():
    """
    Starting spark session
    """

    spark = SparkSession \
        .builder \
        .appName("Python Spark SQL basic example") \
        .config("spark.some.config.option", "some-value") \
        .getOrCreate()

    return spark
spark = start_spark_session()   


def get_imputed_df( IMPUTATION_TRAIN_PATH, IMPUTATION_PREDICT_PATH ):
    """
    Load rack files: genereal preprocessing, preprocessing, imputed
    """
  
    imputed_train_df = load_df( IMPUTATION_TRAIN_PATH )
    imputed_predict_df = load_df( IMPUTATION_PREDICT_PATH )
    imputed_train_df = imputed_train_df.withColumn("normal_2", lit(None)).select("*")
    imputed_predict_df = imputed_predict_df.withColumn("normal", lit(None)).select("*")
    result = imputed_train_df.union(imputed_predict_df)
    
    return result

def load_df( df_path ):
    """
    load any df
    """
    df = spark.read.format("csv").option("header", "true").load( df_path )
    df = df.drop('_c0')
    return df


def suffix_and_join_dfs(df1, df2, on_column ):
    """
    give suffix 1 to first df and suffix 2 to last df
    """
    suffix2 = '2'
    df2 = attach_suffix_to_columns(df2, suffix2)
    on_column2 = on_column + '_' + suffix2
    joined_df = df1.join(df2, col(on_column) == col(on_column2) )
    
    return joined_df


def attach_suffix_to_columns( df, prefix ):    
    """
    """
    columns = df.columns
    df = df.select(*[col(x).alias(x + '_' + prefix) for x in columns])
    return df


def write_to_excel(df, sheet_name):
    """
    How to write to excel
    https://stackoverflow.com/questions/42370977/how-to-save-a-new-sheet-in-an-existing-excel-file-using-pandas
    """
    if 'output.xlsx' not in os.listdir('xls'):
        wb = Workbook()
        wb.save('xls/output.xlsx')
    path = 'xls/output.xlsx' 
    book = load_workbook(path)
    writer = pd.ExcelWriter(path, engine = 'openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    df.toPandas().to_excel(writer, sheet_name = sheet_name)
    writer.save()
    writer.close()

   
def get_module_from_path( module_name , module_path ):

    spec = importlib.util.spec_from_file_location(  module_name , module_path )
    my_module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(my_module)
    
    return my_module
    
def get_te_config_module():


    path = cfg.TE_CONFIG_FILE_PATH
    path_split = os.path.\
        split(path)
    module_name = path_split[1]
    spec = importlib.util.spec_from_file_location(  module_name , path )
    te_config_module = importlib.util.module_from_spec( spec )
    spec.loader.exec_module( te_config_module )
    return te_config_module
     
def get_te_constants_module():


    path = cfg.TE_CONSTANTS_FILE_PATH
    path_split = os.path.\
        split(path)
    module_name = path_split[1]
    spec = importlib.util.spec_from_file_location(  module_name , path )
    te_constants_module = importlib.util.module_from_spec( spec )
    spec.loader.exec_module( te_constants_module )
    
    
    return te_constants_module

def get_te_load_pipeline_config_module():
    sys.path.append(cfg.TE_PATH)
    path = cfg.TE_LOAD_PIPELINE_CONFIG_FILE_PATH
    path_split = os.path.\
        split(path)
    module_name = path_split[1]
    print(path,module_name)
    spec = importlib.util.spec_from_file_location(  module_name , path )
    te_load_pipeline_config_module = importlib.util.module_from_spec( spec )
    spec.loader.exec_module( te_load_pipeline_config_module )
    return te_load_pipeline_config_module


def load_final_train_df():
    with open(cfg.FINAL_TRAIN_PATH, 'rb') as f:
        final_train_df = pickle.load(f)[0] # loading tuple of train and test
    return final_train_df